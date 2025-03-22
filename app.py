import re
import time
import traceback
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

debug = False
excel_file = r"E:\Miscellaneous\Projects\CS\Python\Scripts\Not Github Scripts\card_collection\data\original_data copy.xlsx"

class CardScraper:
    def __init__(self, excel_file, debug=False):
        self.excel_file = excel_file
        self.debug = debug

        try:
            self.wb = load_workbook(self.excel_file)
        except Exception as e:
            raise Exception(f"Error loading workbook: {e}")
        
        try:
            self.cards_sheet = self.wb["Cards"]
        except Exception as e:
            raise Exception(f"Error: 'Cards' sheet not found: {e}")
        
        try:
            self.summary_sheet = self.wb["Summary"]
        except Exception as e:
            raise Exception(f"Error: 'Summary' sheet not found: {e}")
        
        # Determine header positions (assumes header is in row 1)
        header_cells = list(self.cards_sheet.iter_rows(min_row=1, max_row=1, values_only=False))[0]
        self.headers = {}
        for cell in header_cells:
            if cell.value is not None:
                self.headers[cell.value] = cell.column  # 1-indexed
        required_headers = ["Description", "Qty.", "Market Price", "Link"]
        for h in required_headers:
            if h not in self.headers:
                raise Exception(f"Missing required column '{h}' in the Cards sheet.")
        self.url_col = self.headers["Link"]
        self.qty_col = self.headers["Qty."]
        self.desc_col = self.headers["Description"]
        self.market_price_col = self.headers["Market Price"]

        # Set up Selenium Chrome WebDriver in headless mode.
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--no-sandbox")
        chrome_options.add_argument("--disable-dev-shm-usage")
        chrome_options.add_experimental_option("excludeSwitches", ["enable-logging"])
        chrome_options.add_argument("--log-level=3")
        try:
            self.driver = webdriver.Chrome(options=chrome_options)
        except Exception as e:
            raise Exception(f"Error initializing Chrome driver: {e}")

    def process_cards(self):
        for row in self.cards_sheet.iter_rows(min_row=2):
            url_cell = row[self.url_col - 1]
            try:
                if self.debug:
                    print(f"Processing row {url_cell.row}: value='{url_cell.value}', hyperlink={url_cell.hyperlink}")
                # Try to retrieve the URL from the hyperlink attribute first.
                original_url = None
                if url_cell.hyperlink is not None:
                    original_url = url_cell.hyperlink.target
                    if self.debug:
                        print(f"Found hyperlink from cell.hyperlink.target: {original_url}")
                else:
                    original_url = url_cell.value
                    if self.debug:
                        print(f"No hyperlink found; using cell value: {original_url}")

                # If still "Link" or empty, check the internal _hyperlinks.
                if not original_url or original_url == "Link":
                    if self.debug:
                        print("Attempting to check cards_sheet._hyperlinks...")
                    for hl in self.cards_sheet._hyperlinks:
                        if hl.ref == url_cell.coordinate:
                            original_url = hl.target
                            if self.debug:
                                print(f"Found URL in _hyperlinks: {original_url}")
                            break

                if not original_url or original_url == "Link":
                    if self.debug:
                        print(f"No valid URL found for row {url_cell.row}; skipping this row.")
                    continue

                # Clean URL: remove query parameters.
                clean_url = original_url.split('?')[0]
                if self.debug:
                    print(f"Clean URL: {clean_url}")

                try:
                    print("Scraping URL:", clean_url)
                    self.driver.get(clean_url)
                    # Wait for dynamic content to load.
                    time.sleep(3)
                    # Scrape description.
                    try:
                        desc_elem = self.driver.find_element(By.CSS_SELECTOR, ".product-details__name")
                        description = desc_elem.text.strip()
                    except Exception:
                        description = ""
                    # Scrape market price.
                    market_price = 0.0
                    market_price_text = ""
                    try:
                        price_elem = self.driver.find_element(By.CSS_SELECTOR, ".price-points__upper__price")
                        market_price_text = price_elem.text.strip()
                    except Exception:
                        try:
                            price_elem = self.driver.find_element(By.CSS_SELECTOR, ".spotlight__price")
                            market_price_text = price_elem.text.strip()
                        except Exception:
                            market_price_text = ""
                    market_price_clean = re.sub(r"[^\d\.]", "", market_price_text)
                    try:
                        market_price = float(market_price_clean)
                    except Exception:
                        market_price = 0.0
                    if self.debug:
                        print("Scraped data:", description, market_price)
                except Exception as e:
                    if self.debug:
                        print(f"Error scraping {clean_url}: {e}")
                        traceback.print_exc()
                    description, market_price = "", 0.0

                # Update cells.
                row[self.desc_col - 1].value = description
                row[self.market_price_col - 1].value = market_price

                # Set hyperlink first, then update display text.
                url_cell.hyperlink = clean_url
                url_cell.value = "Link"
            except Exception as e:
                if self.debug:
                    print(f"Error processing row {url_cell.row}: {e}")
                    traceback.print_exc()
                continue

    def process_summary(self):
        total_qty = 0
        unique_cards = set()
        total_value = 0
        price_ranges = {"<$1": 0, "$1-$5": 0, "$5-$10": 0, "$10-$50": 0, ">$50": 0}

        cards_data = []
        for row in self.cards_sheet.iter_rows(min_row=2):
            try:
                qty = int(row[self.qty_col - 1].value) if row[self.qty_col - 1].value is not None else 0
            except Exception:
                qty = 0
            total_qty += qty

            try:
                mp = float(row[self.market_price_col - 1].value) if row[self.market_price_col - 1].value is not None else 0.0
            except Exception:
                mp = 0.0
            card_value = qty * mp

            description = row[self.desc_col - 1].value if row[self.desc_col - 1].value is not None else ""
            link = row[self.url_col - 1].hyperlink.target if row[self.url_col - 1].hyperlink is not None else ""

            unique_cards.add(description)
            total_value += card_value

            # Price distribution
            if mp < 1:
                price_ranges["<$1"] += qty
            elif 1 <= mp < 5:
                price_ranges["$1-$5"] += qty
            elif 5 <= mp < 10:
                price_ranges["$5-$10"] += qty
            elif 10 <= mp < 50:
                price_ranges["$10-$50"] += qty
            else:
                price_ranges[">$50"] += qty

            cards_data.append({
                "Description": description,
                "Qty": qty,
                "Market Price": mp,
                "Value": card_value,
                "Link": link
            })

        # Writing Summary
        self.summary_sheet["B2"] = total_qty
        self.summary_sheet["B3"] = len(unique_cards)  # Unique cards count
        self.summary_sheet["B4"] = total_value
        self.summary_sheet["B7"] = price_ranges["<$1"]
        self.summary_sheet["B8"] = price_ranges["$1-$5"]
        self.summary_sheet["B9"] = price_ranges["$5-$10"]
        self.summary_sheet["B10"] = price_ranges["$10-$50"]
        self.summary_sheet["B11"] = price_ranges[">$50"]

        # Determine the top 5 cards by value.
        cards_data_sorted = sorted(cards_data, key=lambda x: x["Market Price"], reverse=True)
        top_5 = cards_data_sorted[:5]
        start_row = 3
        for i, card in enumerate(top_5):
            current_row = start_row + i
            self.summary_sheet.cell(row=current_row, column=4, value=card["Description"])
            self.summary_sheet.cell(row=current_row, column=5, value=card["Qty"])
            self.summary_sheet.cell(row=current_row, column=6, value=card["Market Price"])
            link_cell = self.summary_sheet.cell(row=current_row, column=7, value="Link")
            if card["Link"]:
                link_cell.hyperlink = card["Link"]

    def apply_formatting(self):
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        for row in self.cards_sheet.iter_rows(min_row=2, max_row=self.cards_sheet.max_row):
            desc_cell = row[self.desc_col - 1]
            desc_cell.border = thin_border
            market_price_cell = row[self.market_price_col - 1]
            market_price_cell.border = thin_border
            market_price_cell.alignment = Alignment(horizontal="center")
            market_price_cell.number_format = '"$"#,##0.00'

    def save_workbook(self):
        try:
            self.wb.save(self.excel_file)
            print("Excel file updated successfully.")
        except Exception as e:
            print(f"Error saving workbook: {e}")
            traceback.print_exc()

    def close(self):
        try:
            self.driver.quit()
        except Exception as e:
            if self.debug:
                print(f"Error quitting driver: {e}")

    def run(self):
        self.process_cards()
        self.process_summary()
        self.apply_formatting()
        self.save_workbook()
        self.close()

if __name__ == "__main__":
    scraper = CardScraper(excel_file, debug=debug)
    scraper.run()